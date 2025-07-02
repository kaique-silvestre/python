import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pathlib

ROOT_FOLDER = pathlib.Path(__file__).parent 

WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

students: list[list] = [
    ["Ana Silva", 17, 8.5],
    ["Bruno Costa", 18, 7.2],
    ["Carla Mendes", 16, 9.1],
    ["Diego Rocha", 17, 6.8]
]


# Criando um novo workbook
workbook01 = openpyxl.Workbook()

# Criando uma nova planilha 
worksheet01 = workbook01.active

# Criando o cabeçalho
worksheet01.cell(1, 1, "nome")
worksheet01.cell(1, 2, "idade")
worksheet01.cell(1, 3, "nota")

for index01, student_row in enumerate(students, start=2):
    for index02, student_column in enumerate(student_row, start=1):
        worksheet01.cell(index01, index02, student_column)

# Salvando o workbook em um caminho 
workbook01.save(WORKBOOK_PATH)



