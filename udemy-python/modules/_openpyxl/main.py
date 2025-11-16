import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pathlib

ROOT_FOLDER = pathlib.Path(__file__).parent 

WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

students: list[list] = [
    ["Ana Silva", 17, 8.5],
    ["Bruno Costa", 18, 7.2],
    ["Carla Mendes", 16, 9.1],
    ["Diego Rocha", 17, 6.8]
]


# Criando um novo workbook
workbook01 = openpyxl.Workbook()

# Ativa a planilha padrão do workbook
worksheet01 = workbook01.active

workbook01.create_sheet("Teste", 0) # Cria uma nova planilha chamada "Teste" na posição 0

worksheet01 = workbook01["Teste"]  # Acessa a planilha "Teste"

#workbook01.remove(worksheet01)  # Remove a planilha "Teste"



print(workbook01.sheetnames)  # Exibe os nomes das planilhas no workbook

# Criando o cabeçalho
worksheet01.cell(1, 1, "nome")
worksheet01.cell(1, 2, "idade")
worksheet01.cell(1, 3, "nota")

for index01, student_row in enumerate(students, start=2):
    for index02, student_column in enumerate(student_row, start=1):
        worksheet01.cell(index01, index02, student_column)

for student in students:
    worksheet01.append(student)



# Salvando o workbook em um caminho 
workbook01.save(WORKBOOK_PATH)



